import os
import logging
import functools
import copy
from datetime import datetime
# need to import _strptime for multi-threading, a known python 2.7 bug
import _strptime
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font
import pyani.core.appvars
import pyani.core.anivars
import pyani.core.ui
import pyani.core.util
from pyani.core.mngr.core import AniCoreMngr


logger = logging.getLogger()


class AniAssetMngr(AniCoreMngr):
    """
    A class object that manages asset information.

    Requirements:
        - show assest are in /LongGong/assets/
        - the folders beneath the above assets/ folder are called asset types, ex: /LongGong/assets/char
        - the folders beneath the asset types are the asset names, ex: /LongGong/assets/char/charAnglerFish
        - which each asset's folder are asset components, ex: /LongGong/assets/char/charAnglerFish/rigs
        - asset files are either in approved/ or work/ folders,
          for example, /LongGong/assets/char/charAnglerFish/rigs/approved or
          /LongGong/assets/char/charAnglerFish/rigs/work
        - approved asset files have no version in file name and versions live in approved/history
        - work files have versions in file names and all work files live beneath the work/ folder, i.e. no history
          folder like approved
        - notes live directly under approved/history or work/ and have version name in them
        - shot assets live in /LongGong/sequence/Seq###/Shot###/

    Methods prefixed with 'server' are methods that contact CGT's cloud.

    CACHE FORMAT:
    Creates a local representation of the CGT server file structure and information about the assets.
    The cache data file name is in the cgt_asset_info_cache_path member variable in pyani.core.appvars.AppVars class.
    The cache is located in the permanent data directory - see pyani.core.appvars.
    The cache is a dictionary in the following format:
            {
            "asset type" - this is char, set, prop, shot
                "asset category/component": { - rig, gpu cache, audio
                    "asset name": { - actual name of the asset, such as charHei
                        metadata as metadata_name: value, value can be any type such as list or string. Contains
                        things like version, files associated with the asset, etc...
                    }, ...
                }, ...
        }

    supported metadata is:
            {
                "local path": string path on the local machine of the directory holding files,
                "files": list of strings representing file names (not full path, just file name,
                "cgt cloud dir": string path on the server of the directory holding files,
                "version": string version,
                "approved": boolean,
                "notes path": string path on server
            }


    USAGE:

    Create one instance to handle multiple tools. Set the active component to control what assets are being managed
    EX: We have Rigs and Audio
    When I want to manage Rigs, I do asset_mngr_instance.active_asset_component = "Rigs" - set to user friendly name
    see pyani.core.appvars.AppVars asset types 'name' for user friendly name

    Updating / Building Caches:
    To build a local cache,
    simply call sync_local_cache_with_server() (optionally pass a dict to update specific assets, see method
    doc string for format) and connect to the signal _thread_server_sync_complete (inherited from AniCoreMngr)
    to find out when it finishes.

    Download Files and Sync Cache:
    call sync_local_cache_with_server_and_download and pass a dict to update, see method
    doc string for format). connect to _thread_server_sync_complete (inherited from AniCoreMngr)
    to find out when it finishes.

    Reports errors using signals from pyani.core.mngrcore.AniCoreMngr:
    error_thread_signal, connect to this signal as:
    asset_mngr_instance.error_thread_signal.connect(method_that_receives_error)

    method_that_receives_error(error):
        print error

    emits these, and listening class can receive and display.

    """

    def __init__(self):
        AniCoreMngr.__init__(self)
        self._asset_info = dict()

        # identifies which component this mngr is currently responsible for
        self.active_asset_component = None

        # records of changed audio stored as dict:
        '''
        { 
            seq name: [ 
                (shot name, modified date), 
                (shot name, modified date),
                ...
            ] 
        }
        '''
        self.shots_with_changed_audio = dict()

        # record of errors that occur checking audio timestamps  as dict
        # { seq name: {shot name: error}, {shot name: error}, ... }
        self.shots_failed_checking_timestamp = dict()

        # list of existing assets so we can compare after a sync for newly added or updated assets
        self._existing_assets_before_sync = dict()
        # assets timestamp before downloads
        self._assets_timestamp_before_dl = dict()

    @property
    def active_asset_component(self):
        return self._active_asset_component

    @active_asset_component.setter
    def active_asset_component(self, component_name):
        self._active_asset_component = component_name

    def get_asset_server_dir_from_cache(self, asset_type, asset_component, asset_name):
        """
        Access method to get server path (the directory holding the files for the asset). Allows dict to change format
        and only need to change here
        :param asset_type: the asset type - see pyani.core.appvars.py for asset components
        :param asset_component: the asset component - see pyani.core.appvars.py for asset components
        :param asset_name: the name of the asset as a string
        :return: a string path on the server
        """
        return self._asset_info[asset_type][asset_component][asset_name]["cgt cloud dir"]

    def get_asset_local_dir_from_cache(self, asset_type, asset_component, asset_name):
        """
        Access method to get local path (the directory holding the files for the asset). Allows dict to change format
        and only need to change here
        :param asset_type: the asset type - see pyani.core.appvars.py for asset components
        :param asset_component: the asset component - see pyani.core.appvars.py for asset components
        :param asset_name: the name of the asset as a string
        :return: a string path on the local machine
        """
        return self._asset_info[asset_type][asset_component][asset_name]["local path"]

    def get_asset_version_from_cache(self, asset_type, asset_component, asset_name):
        """
        Access method to get asset version. Allows dict to change format
        and only need to change here
        :param asset_type: the asset type - see pyani.core.appvars.py for asset components
        :param asset_component: the asset component - see pyani.core.appvars.py for asset components
        :param asset_name: the name of the asset as a string
        :return: a string representing the version
        """
        return self._asset_info[asset_type][asset_component][asset_name]["version"]

    def get_asset_files(self, asset_type, asset_component, asset_name):
        """
        Access method to get the asset file names. Allows dict to change format
        and only need to change here
        :param asset_type: the asset type - see pyani.core.appvars.py for asset components
        :param asset_component: the asset component - see pyani.core.appvars.py for asset components
        :param asset_name: the name of the asset as a string
        :return: a list of file names (only the file name, not the full path)
        """
        return self._asset_info[asset_type][asset_component][asset_name]["files"]

    def check_for_new_assets(self, asset_component, asset_list=None):
        """
        Checks for assets that have changed since last run.
        :param asset_component: the asset component - see pyani.core.appvars.py for asset components
        :param asset_list: optional list of assets to check
        :return:
        """
        if asset_component == "audio":
            self._check_for_new_audio(seqs=asset_list)

    def load_server_asset_info_cache(self):
        """
        reads the server asset info cache off disk
        :return: None if the file was read successfully, the error as a string if reading is unsuccessful.
        """
        json_data = self.load_server_local_cache(self.app_vars.cgt_asset_info_cache_path)
        if isinstance(json_data, dict):
            self._asset_info = json_data
            return None
        else:
            return json_data

    def get_asset_component_names(self):
        """
        Gets a list of all possible asset components using user friendly name - see pyani.core.appvars.AppVars
        :return: list of asset components
        """
        components = []
        for asset_type in self.app_vars.asset_types:
            for component in self.app_vars.asset_types[asset_type]:
                if self.app_vars.asset_types[asset_type][component]['name'] not in components:
                    components.append(self.app_vars.asset_types[asset_type][component]['name'])
        return components

    def get_asset_type_by_asset_component_name(self, asset_component):
        """
        Given an asset component, finds all asset types that have this component
        :param asset_component: the asset component - see pyani.core.appvars.py for asset components
        :return: the asset type(s) as a list, or None if no asset types were found
        """
        asset_types_list = []
        for asset_type in self._asset_info:
            if asset_component in self._asset_info[asset_type]:
                asset_types_list.append(asset_type)
        return asset_types_list

    def get_asset_info_by_asset_name(self, asset_type, asset_component, asset_name):
        """
        Gets all assets and their info given a asset_component name
        :param asset_type: the asset type - see pyani.core.appvars.py for asset components
        :param asset_component: the asset component - see pyani.core.appvars.py for asset components
        :param asset_name: the name of the asset as a string
        :return: a tuple in format:
        {
            asset type,
            asset component,
            asset name,
            { asset properties - see class docstring for supported properties}
        }

        or returns none if the asset_type or asset_component or asset name doesn't exist
        """
        try:
            asset_info = (
                asset_type,
                asset_component,
                asset_name,
                self._asset_info[asset_type][asset_component][asset_name]
            )
            return asset_info
        except KeyError:
            return None

    def get_assets_by_asset_component(self, asset_type, asset_component):
        """
        Gets all assets given an asset_component name
        :param asset_type: the asset type - see pyani.core.appvars.py for asset components
        :param asset_component: the asset component - see pyani.core.appvars.py for asset components
        :return: a list of asset names or returns none if the asset_type or asset_component doesn't exist
        """
        try:
            return self._asset_info[asset_type][asset_component].keys()
        except KeyError:
            return None

    def get_release_notes(self, asset_component, asset_name):
        """
        Gets the release notes for the asset from CGT
        :param asset_component: the asset component - see pyani.core.appvars.py for asset components
        :param asset_name: the asset name as a string
        :return: a tuple of the notes (string) and error if any (string).
        """
        # the python script to call that connects to cgt
        py_script = os.path.join(self.app_vars.cgt_bridge_api_path, "cgt_get_notes.py")

        # capitalize the component first letter for CGT
        asset_component = asset_component[0].upper() + asset_component[1:]

        # the command that subprocess will execute
        command = [
            py_script,
            asset_component,
            asset_name,
            self.app_vars.cgt_ip,
            self.app_vars.cgt_user,
            self.app_vars.cgt_pass,
        ]

        try:
            output, error = pyani.core.util.call_ext_py_api(command)

            # check for subprocess errors
            if error:
                error_fmt = "Error occurred launching subprocess. Error is {0}".format(error)
                self.send_thread_error(error_fmt)
                return None, error_fmt

            # check for output and return
            if output:
                return output, None

        # CGT errors
        except pyani.core.util.CGTError as error:
            error_fmt = "Error occurred connecting to CGT. Error is {0}".format(error)
            self.send_thread_error(error_fmt)
            return None, error_fmt

    def is_asset_publishable(self, asset_type, asset_component):
        """
        checks if an asset is publishable, meaning it has an approved folder and possibly a work folder
        :param asset_type: the type of asset - see pyani.core.appvars.py for asset types
        :param asset_component: the asset component to check - see pyani.core.appvars.py for asset components
        :return: True if it can be published, False if not
        """
        if self.app_vars.asset_types[asset_type][asset_component]['is publishable']:
            return True
        else:
            return False

    def is_asset_approved(self, asset_type, asset_component, asset_name):
        """
        checks if an asset is approved, meaning it has an approved folder
        :param asset_type: the type of asset - see pyani.core.appvars.py for asset types
        :param asset_component: the asset component to check - see pyani.core.appvars.py for asset components
        :param asset_name: the name of the asset
        :return: True if it is approved, false if not
        """
        # load asset info if not loaded
        if not self._asset_info:
            error = self.load_server_asset_info_cache()
            if error:
                error_fmt = "Problem loading asset info cache in assets.py - is_asset_approved. Error is {0}".format(
                    error
                )
                logger.error(error_fmt)
                return False
        if self._asset_info[asset_type][asset_component][asset_name]['approved']:
            return True
        else:
            return False

    def is_asset_versioned(self, asset_type, asset_component):
        """
        checks if an asset is versioned
        :param asset_type: the type of asset - see pyani.core.appvars.py for asset types
        :param asset_component: the asset component to check - see pyani.core.appvars.py for asset components
        :return: True if it is versioned, False if not
        """
        if self.app_vars.asset_types[asset_type][asset_component]['is versioned']:
            return True
        else:
            return False

    def asset_component_supports_release_notes(self, asset_component):
        """
        checks if an asset's component can have release notes
        :param asset_component: the asset component to check - see pyani.core.appvars.py for asset components
        :return: True if it can have release notes, False if not:
        """
        for asset_type in self.app_vars.asset_types:
            if asset_component in self.app_vars.asset_types[asset_type]:
                # doesn't matter which asset type the component belongs to. components have the same properties
                # regardless of asset type. For example, set rigs and char rigs both support notes, because rigs
                # which is an asset component supports notes
                if self.app_vars.asset_types[asset_type][asset_component]['supports notes']:
                    return True
        return False

    def update_config_file_by_component_name(self, selected_asset_component, config_data):
        """
        Updates the asset update config file with new assets and removes assets that are de-selected (not in
        selection)
        :param selected_asset_component: the asset component we are updating
        :param config_data: a dict of the asset type, asset component and asset names we want to add to the config
        :return: error if occurs, otherwise None
        """
        # if the config file doesn't exist, just save the data
        if not os.path.exists(self.app_vars.update_config_file):
            if not os.path.exists(self.app_vars.persistent_data_path):
                error = pyani.core.util.make_dir(self.app_vars.persistent_data_path)
                if error:
                    return error
            error = pyani.core.util.write_json(
                self.app_vars.update_config_file,
                config_data,
                indent=4
            )
            if error:
                return error
            return None
        # file exists
        else:
            # pull the config data off disk
            existing_config_data = pyani.core.util.load_json(self.app_vars.update_config_file)
            # check if config data is an empty file, if so set to a empty dict object
            if not isinstance(existing_config_data, dict):
                existing_config_data = dict()

            # first check for assets whose type and component don't exist yet in the config file
            for asset_type in config_data:
                # when the asset type doesn't yet exist, but other asset types do in the file, so can just add.
                if asset_type not in existing_config_data:
                    existing_config_data[asset_type] = dict()
                    existing_config_data[asset_type] = config_data[asset_type]

                # when asset component doesn't exist but the type does, so can just add
                elif selected_asset_component not in existing_config_data[asset_type]:
                    existing_config_data[asset_type][selected_asset_component] = dict()
                    existing_config_data[asset_type][selected_asset_component] = \
                        config_data[asset_type][selected_asset_component]
                # just replace since type and component exist
                else:
                    existing_config_data[asset_type][selected_asset_component] = \
                        config_data[asset_type][selected_asset_component]

            # check for assets that got deselected for the active component
            for existing_asset_type in existing_config_data:
                for existing_asset_component in existing_config_data[existing_asset_type]:
                    # check if the component from the existing config is in the selection. Otherwise it will delete
                    # assets from other components. For example if we pass only audio (i.e. on audio tab), it would
                    # delete rigs. This check prevents that
                    if selected_asset_component == existing_asset_component:
                        # if don't find the key, means doesn't exist in the selected assets passed in, so remove from
                        # our existing config file
                        if not pyani.core.util.find_val_in_nested_dict(
                                config_data,
                                [existing_asset_type, existing_asset_component]
                        ):
                            existing_config_data[existing_asset_type][existing_asset_component] = dict()

            error = pyani.core.util.write_json(self.app_vars.update_config_file, existing_config_data, indent=4)
            if error:
                return error
            return None

    def update_config_file_after_sync(self, debug=False):
        # pull the config data off disk
        existing_config_data = pyani.core.util.load_json(self.app_vars.update_config_file)
        # check if config data loaded
        if not isinstance(existing_config_data, dict):
            error = "Error loading update config file from disk. Error is: {0}".format(existing_config_data)
            self.send_thread_error(error)
            return error

        # load asset info cache
        error = self.load_server_asset_info_cache()
        if error:
            self.send_thread_error(error)
            return error

        # assets in current update config
        assets_in_update_config = {key: value for key, value in existing_config_data.items() if not key == 'tools'}

        # assets to remove - list of tuples where tuple is (asset type, asset component, asset name)
        assets_to_remove = list()

        # find any assets that no longer exist
        if assets_in_update_config:
            for asset_type in assets_in_update_config:
                for asset_component in assets_in_update_config[asset_type]:
                    for asset_name in assets_in_update_config[asset_type][asset_component]:
                        if not pyani.core.util.find_val_in_nested_dict(
                                self._asset_info, [asset_type, asset_component, asset_name]
                        ):
                            assets_to_remove.append((asset_type, asset_component, asset_name))

        # remove non-existent assets, also removes any empty asset components and asset types
        if assets_to_remove:
            for asset in assets_to_remove:
                asset_type, asset_component, asset_name = asset
                # remove asset
                existing_config_data[asset_type][asset_component].remove(asset_name)
                # check if asset_component empty
                if not existing_config_data[asset_type][asset_component]:
                    del existing_config_data[asset_type][asset_component]
                # check if asset type empty
                if not existing_config_data[asset_type]:
                    del existing_config_data[asset_type]

        if debug:
            print "Updated Config Data Is Now:"
            print existing_config_data
        else:
            error = pyani.core.util.write_json(self.app_vars.update_config_file, existing_config_data, indent=4)
            if error:
                error_fmt = "Could not save sync'd update config file. Error is {0}".format(error)
                self.send_thread_error(error_fmt)
                return error_fmt

        self.finished_signal.emit(None)
        return None

    def update_local_cache(self, asset_type, asset_component, asset_name, asset_update_info):
        """
        Updates the cache on disk with data provided. Does not connect to server. Only uses the asset info passed in
        through asset_update_info
        :param asset_type: a list of the type of asset(s) to update - see pyani.core.appvars.py for asset types
        :param asset_component: a list of the asset component to update- see pyani.core.appvars.py for asset components
        :param asset_name: a list of asset names
        :param asset_update_info: a tuple in format:
            ( asset type as string, asset component as string, asset name as string, asset info as dict )
            asset info format is
            {
                "local path": string path on the local machine of the directory holding files,
                "files": list of strings representing file names (not full path, just file name,
                "cgt cloud dir": string path on the server of the directory holding files,
                "version": string version,
                "approved": boolean,
                "notes path": string path on server
            }

            one or more of the above can be provided, for example these are acceptable:
            {
                "local path": string path,
                "files": list of string file names
            }
            -or-
            {
                 "local path": string path
            }
            -or-
            {
                "files": list of string file names,
                "cgt cloud dir": string path,
                "version": string version,
                "approved": boolean,
                "notes path": string path
            }
            etc...
        :return: None if updated cache, an error string if couldn't update. Note for an entire cache rebuild, use
        the signal finished to check for errors, since its multi-threaded
        """
        try:
            # update the asset info based off the provided data
            for key, value in asset_update_info.items():
                self._asset_info[asset_type][asset_component][asset_name][key] = value
        except (KeyError, ValueError) as e:
            return "Could not update the local cache. Error is {0}".format(e)

    def sync_local_cache_with_server_and_download_gui(self, update_data_dict):
        """
        used with gui asset mngr
        Updates the cache on disk with the current server data. If no parameters are filled the entire cache will
        be rebuilt.
        :param update_data_dict: a dict in format:
        {
         asset type: {
             asset component(s): [
                 asset name(s)
             ]
             }, more asset types...
        }

        There can be one or more asset types. Asset components and asset names are optional.
        Asset components require an asset type. Asset Names require both an asset type and asset component.
        a list of the type of asset(s) to update - see pyani.core.appvars.py for asset types and asset components

        :return: None if updated cache, an error string if couldn't update. Note for an entire cache rebuild, use
        the signal finished to check for errors, since its multi-threaded
        """
        # no asset types, so can't set any other values in data struct, so rebuild entire cache
        if not update_data_dict:
            return "At least one asset must be provided to update."

        # reset progress
        self.init_progress_window("Sync Progress", "Updating Assets...")

        # update the local cache for the assets given
        self.server_build_local_cache(
            assets_dict=update_data_dict,
            thread_callback=self._thread_server_sync_complete,
            thread_callback_args=[self.active_asset_component, self.server_save_local_cache]
        )

        # downloads from the gui, which requires knowing which asset component was run
        self.server_download(update_data_dict, gui_mode=True)

    def sync_local_cache_with_server(self, update_data_dict=None):
        """
        Updates the cache on disk with the current server data. If no parameters are filled the entire cache will
        be rebuilt.
        :param update_data_dict: a dict in format:
        {
            asset type: {
                asset component(s): [
                    asset name(s)
                ]
                }, more asset types...
        }

        There can be one or more asset types. Asset components and asset names are optional.
        Asset components require an asset type. Asset Names require both an asset type and asset component.
        a list of the type of asset(s) to update - see pyani.core.appvars.py for asset types and asset components

        :return: None if updated cache, an error string if couldn't update. Note for an entire cache rebuild, use
        the signal finished to check for errors, since its multi-threaded
        """

        # reset threads counters
        self.thread_total = 0.0
        self.threads_done = 0.0

        # no asset types, so can't set any other values in data struct, so rebuild entire cache
        if not update_data_dict:
            self.server_build_local_cache()
        else:
            self.server_build_local_cache(
                assets_dict=update_data_dict,
                thread_callback=self._thread_server_sync_complete,
                thread_callback_args=[self.active_asset_component, self.server_save_local_cache]
            )

    def server_download(self, assets_dict=None, gui_mode=False):
        """
        downloads files. If an asset list is provided only those assets will be downloaded, otherwise all assets are
        downloaded. Gui mode provides cache syncing, otherwise the local cgt cache is not synced during download.
        Uses multi-threading.
        :param assets_dict: a dict in format:
        {
             asset type: {
                 asset component(s): [
                     asset name(s)
                 ]
                 }, more asset types...
        }
        :param gui_mode: if True connects a slot that sends the active tool tab name and saves the local tool cache
                         for use in gui mode of asset mngr
        """
        # set number of threads to max - can do this since running per asset
        self.set_number_of_concurrent_threads()

        # if not in gui mode reset thread count and errors, otherwise don't because cache sync did this already
        if not gui_mode:
            # reset thread counters
            self._reset_thread_counters()
            # reset error list
            self.init_thread_error()

        # if not visible then no other function called this, so we can show progress window
        if not self.progress_win.isVisible():
            # reset progress
            self.init_progress_window("Sync Progress", "Updating assets...")

        # make sure asset info is loaded
        if not self._asset_info:
            error = self.load_server_asset_info_cache()
            if error:
                self.send_thread_error(error)
                return error

        # check if assets to download were provided, if not download all assets
        if not assets_dict:
            assets_dict = self._asset_info

        # now use multi-threading to download
        for asset_type in assets_dict:
            for asset_component in assets_dict[asset_type]:
                for asset_name in assets_dict[asset_type][asset_component]:
                    # make sure asset exists, can't download non-existent asset
                    if pyani.core.util.find_val_in_nested_dict(
                            self._asset_info,
                            [asset_type, asset_component, asset_name]
                    ):
                        # possible no files, then skip, otherwise download
                        file_names = self.get_asset_files(asset_type, asset_component, asset_name)
                        if file_names:
                            # could be more than one file
                            for file_name in file_names:
                                local_path = self.get_asset_local_dir_from_cache(asset_type, asset_component, asset_name)

                                # get timestamps of tools being downloaded - create keys if needed
                                if asset_type not in self._assets_timestamp_before_dl:
                                    self._assets_timestamp_before_dl[asset_type] = dict()
                                if asset_component not in self._assets_timestamp_before_dl[asset_type]:
                                    self._assets_timestamp_before_dl[asset_type][asset_component] = dict()
                                if asset_name not in self._assets_timestamp_before_dl[asset_type][asset_component]:
                                    self._assets_timestamp_before_dl[asset_type][asset_component][asset_name] = dict()
                                file_path = "{0}\\{1}".format(local_path, file_name.split("/")[-1])
                                # file may not be on local machine, so try to get time, if can't set to 0
                                try:
                                    self._assets_timestamp_before_dl[asset_type][asset_component][asset_name][file_path] = os.path.getmtime(file_path)
                                except WindowsError:
                                    self._assets_timestamp_before_dl[asset_type][asset_component][asset_name][file_path] = 0.0

                                # server_file_download expects a list of files, so pass list even though just one file
                                worker = pyani.core.ui.Worker(
                                    self.server_file_download,
                                    False,
                                    [file_name],
                                    local_file_paths=[local_path],
                                    update_local_version=True
                                )
                                self.thread_total += 1.0
                                self.thread_pool.start(worker)

                                # slot that is called when a thread finishes
                                if gui_mode:
                                    # passes the active component so calling classes can know what was updated
                                    # and the save cache method so that when cache gets updated it can be saved
                                    worker.signals.finished.connect(
                                        functools.partial(
                                            self._thread_server_sync_complete,
                                            self.active_asset_component,
                                            self.server_save_local_cache
                                        )
                                    )
                                else:
                                    worker.signals.finished.connect(self._thread_server_download_complete)

                                worker.signals.error.connect(self.send_thread_error)

    def server_build_local_cache(self, assets_dict=None, thread_callback=None, thread_callback_args=None):
        """
        Creates a asset data struct using server data. Uses multi-threading to gather data and store it locally.
        Stored in the persistent data location. This cache has version info and paths so apps don't have to
        access server for this info. Uses multi-threading. Either builds entire cache or updates based off a assets
        in the asset_dict parameter.
        :param assets_dict: a dict in format:
        {
            asset type: {
                asset component(s): [
                    asset name(s)
                ]
                }, more asset types...
        }
         corresponding to the asset type in asset types list. ex:
        :param thread_callback: a method to call as threads complete
        :param thread_callback_args: any args to pass to thread callback
        """
        # set number of threads to max
        self.set_number_of_concurrent_threads()

        self._reset_thread_counters()

        # if no thread callback then normal server cache creation so show progress, otherwise there should be
        # a progress window already running
        if not thread_callback:
            # reset progress
            self.init_progress_window("Cache Progress", "Creating cache...")

        # load existing cache if exists and store in a copy
        error = self.load_server_asset_info_cache()
        if not error:
            self._existing_assets_before_sync = copy.deepcopy(self._asset_info)

        # if no asset type was provided, rebuild cache for all asset types
        if not assets_dict:
            asset_types = self.app_vars.asset_types
            # reset cache, doing a complete rebuild
            self._asset_info = dict()
        else:
            asset_types = assets_dict.keys()

        for asset_type in asset_types:
            if asset_type not in self._asset_info:
                self._asset_info[asset_type] = dict()
            # if no asset components were provided, rebuild for all components, otherwise rebuild for the
            # asset type's component that was provided
            if not pyani.core.util.find_val_in_nested_dict(assets_dict, [asset_type]):
                asset_components = self.app_vars.asset_types[asset_type].keys()
            else:
                # get asset components provided
                asset_components = assets_dict[asset_type].keys()

            # build path to asset types directory, ex /LongGong/asset/set
            # do this here because all asset components share the same asset type root path, for example rig and
            # model cache for sets are both /LongGong/asset/set
            asset_type_root_path = self.app_vars.asset_types[asset_type][asset_components[0]]["root path"]

            for asset_component in asset_components:
                if asset_component not in self._asset_info[asset_type]:
                    self._asset_info[asset_type][asset_component] = dict()

                # if asset names provided, get asset names
                if pyani.core.util.find_val_in_nested_dict(assets_dict, [asset_type, asset_component]):
                    # grab the assets corresponding to the asset type
                    asset_names = assets_dict[asset_type][asset_component]
                else:
                    asset_names = None

                # now use multi-threading to get file info for assets by component type
                worker = pyani.core.ui.Worker(
                    self.server_get_asset_info,
                    False,
                    asset_type_root_path,
                    asset_type,
                    asset_component,
                    asset_names=asset_names
                )
                self.thread_total += 1.0
                self.thread_pool.start(worker)
                # reset error list
                self.init_thread_error()
                # slot that is called when a thread finishes, pass the call back function to call when its done
                # check if thread callback is cache update or cache update with download, if no callback,
                # use the default cache complete callback
                if not thread_callback:
                    worker.signals.finished.connect(
                        functools.partial(self._thread_server_cache_complete, self.server_save_local_cache)
                    )
                else:
                    active_asset_component = thread_callback_args[0]
                    save_method = thread_callback_args[1]
                    worker.signals.finished.connect(
                        functools.partial(self._thread_server_sync_complete, active_asset_component, save_method)
                    )
                worker.signals.error.connect(self.error_thread_signal)

    def server_get_asset_info(self, root_path, asset_type, asset_component, asset_names=None):
        """
        gets file info for assets by component type from cgt server and adds to asset info cache in permanent dir
        :param root_path: the path to the asset names, for example /LongGong/asset/set/
        :param asset_type: the type of asset, see pyani.core.appvars for asset types
        :param asset_component: the asset component, see pyani.core.appvars for asset components
        :param asset_names: list of asset names
        :return error message or none
        """
        # the file name to store file info from CGT
        if asset_component.split("/") > 1:
            prefix = asset_type + "_" + asset_component.replace("/", "_")
        else:
            prefix = asset_type + "_" + asset_component
        json_temp_file_info_path = os.path.join(
            self.app_vars.cgt_temp_file_cache_dir,
            prefix + "_" + self.app_vars.cgt_tmp_file_cache_filename
        )

        # get file info for assets from CGT
        error = self.server_get_file_listing_using_folder_filter(root_path, asset_component, json_temp_file_info_path)
        if error:
            error_fmt = "Error getting file information from cgt server. Error is {0}".format(error)
            self.send_thread_error(error_fmt)
            return error_fmt

        # process and add cgt file info for assets to asset info cache
        self._create_asset_info_cache(
            root_path, json_temp_file_info_path, asset_type, asset_component, asset_names=asset_names
        )

        return None

    def _create_asset_info_cache(self, root_path, json_temp_file_info_path, asset_type, asset_component, asset_names=None):
        """
        creates the asset info cache stored in permanent data dir
        :param root_path: the path to the asset names, for example /LongGong/asset/set/
        :param json_temp_file_info_path: path where the server file info is stored
        :param asset_type: the type of asset, see pyani.core.appvars for asset types
        :param asset_component: the asset component, see pyani.core.appvars for asset components
        :param asset_names: list of asset names
        """

        asset_info_sorted = self._convert_cgt_file_info_to_asset_info(
            root_path, json_temp_file_info_path, asset_type, asset_component, asset_names=asset_names
        )

        # go through all folders under the root path
        for asset_name in asset_info_sorted:

            # check if the asset exists already, if not add the asset name key and create file list key for asset
            if asset_name not in self._asset_info[asset_type][asset_component]:
                self._asset_info[asset_type][asset_component][asset_name] = dict()

            # check if asset has files in approved
            if 'approved' in asset_info_sorted[asset_name]:
                self._asset_info[asset_type][asset_component][asset_name]["approved"] = True
                # get directory - all files in same directory so just use first file but make sure actually has files
                if asset_info_sorted[asset_name]['approved']:
                    cgt_dir = asset_info_sorted[asset_name]['approved'][0].split('approved')[0] + "approved"
                else:
                    cgt_dir = asset_info_sorted[asset_name]['component path'] + "approved"
                # get files
                file_list = asset_info_sorted[asset_name]['approved']
                # check if the asset is versioned. A publishable asset may not be versioned, like audio
                if self.is_asset_versioned(asset_type, asset_component):
                    # make sure there is a history
                    if 'approved/history' in asset_info_sorted[asset_name]:
                        _, version = self.core_get_latest_version(
                            file_list=asset_info_sorted[asset_name]['approved/history']
                        )
                    else:
                        version = ""
                # asset not versioned but is approved - ex: audio files
                else:
                    version = ""

            # check if asset has a work folder
            elif 'work' in asset_info_sorted[asset_name]:
                self._asset_info[asset_type][asset_component][asset_name]["approved"] = False
                # get directory - all files in same directory so just use first file but make sure actually has files
                if asset_info_sorted[asset_name]['work']:
                    cgt_dir = asset_info_sorted[asset_name]['work'][0].split('work')[0] + "work"
                else:
                    cgt_dir = asset_info_sorted[asset_name]['component path'] + "work"

                # get version first, then can grab the file we want
                _, version = self.core_get_latest_version(file_list=asset_info_sorted[asset_name]['work'])
                # filter for the files that have the version we want
                file_list = [file_name for file_name in asset_info_sorted[asset_name]['work'] if version in file_name]

            # check if asset is not publishable
            elif '.' in asset_info_sorted[asset_name]:
                # get directory - all files in same directory so just use first file
                cgt_dir = '/'.join(asset_info_sorted[asset_name]['.'][0].split('/')[:-1])
                # these aren't versioned
                version = ""
                file_list = asset_info_sorted[asset_name]['.']
            # asset doesn't have files
            else:
                # get directory of component since files, remove right most '/' to be consistent with other asset
                # cgt dir paths
                cgt_dir = asset_info_sorted[asset_name]['component path'].rstrip('/')
                # these aren't versioned
                version = ""
                # no files
                file_list = list()

            self._asset_info[asset_type][asset_component][asset_name]["cgt cloud dir"] = cgt_dir
            self._asset_info[asset_type][asset_component][asset_name]["local path"] = \
                self.convert_server_path_to_local_server_representation(cgt_dir)
            # save the version and file name
            self._asset_info[asset_type][asset_component][asset_name]["version"] = version
            self._asset_info[asset_type][asset_component][asset_name]["files"] = file_list

        return None

    def server_save_local_cache(self):
        """
        Saves the server asset info to a json file
        :return: None if the file was written successfully, the error as a string if writing is unsuccessful.
        If this function is called in a threaded environment, connect to the
        pyani.core.mngrcore thread error signal to get the error
        """
        # check if folder exists, if not make it
        if not os.path.exists(self.app_vars.persistent_data_path):
            error = pyani.core.util.make_dir(self.app_vars.persistent_data_path)
            if error:
                error_fmt = "Could not save local assets cache. Error is {0}".format(error)
                self.send_thread_error(error_fmt)
                return error_fmt

        # creates or overwrites the existing server asset cache json file
        error = pyani.core.util.write_json(self.app_vars.cgt_asset_info_cache_path, self._asset_info, indent=4)
        if error:
            error_fmt = "Could not save local assets cache. Error is {0}".format(error)
            self.send_thread_error(error_fmt)
            return error_fmt
        else:
            return None

    def find_changed_assets(self):
        """
        Passes member variables to parent class function
        :return: a dictionary of assets added, a dictionary of assets modified, a dictionary of assets removed. All
        dictionaries are in the format:

        format:
        {
            asset_type: {
                asset_category: [
                    list of assets
                ]
            }
        }
        """
        return self.find_new_and_updated_assets(
            self._assets_timestamp_before_dl, self._existing_assets_before_sync, self._asset_info
        )

    def _convert_cgt_file_info_to_asset_info(
            self,
            root_path,
            json_temp_file_info_path,
            asset_type,
            asset_component,
            asset_names=None
    ):
        """
        takes a list of files, where each list item is a dict that contains file information such as path
        and modified date and converts them to a list of assets, where each asset is a dict that contains the
        asset's files and other information.
        For example, we go from:
        [
            {/LongGong/asset/set/setAltar/model/cache/cache1.gpu, some other file stats},
            {/LongGong/asset/set/setAltar/model/cache/cache2.gpu, some other file stats},
        ]
         to

        {
            setAltar:
            {
                '.' : [
                    /LongGong/asset/set/setAltar/model/cache/cache1.gpu,
                    /LongGong/asset/set/setAltar/model/cache/cache2.gpu
                ]
            }
        }

        format is:
        {
            asset name:
            {
                'approved' : [file names as a list of strings]
                'approved/history : [file names as a list of strings]
                'work' : [file names as a list of strings]
                '.' :  [file names as a list of strings]
                'component path' : string of the server path to the component, like
                /LongGong/asset/set/setAltar/model/cache
                'modified date': last date modified as string in format yyyy-mm-dd hh:mm:ss
            },
            more assets....
        }

        Published assets will have one or all of the keys: approved, approved/history, work. Non-published assets will
        have '.'
        :param root_path: the path to the asset names, for example /LongGong/asset/set/
        :param json_temp_file_info_path: path where the server file info is stored
        :param asset_type: the asset type - see pyani.core.appvars.py for asset components
        :param asset_component: the asset component - see pyani.core.appvars.py for asset components
        :param asset_names: optional list of asset names to update rather than update all assets for a given asset type
        and component
        :return: a dict of asset names and their associated files
        """

        # load the asset info retrieved from the server
        files_in_path = pyani.core.util.load_json(json_temp_file_info_path)
        if not isinstance(files_in_path, list):
            error_fmt = "Error loading temp cgt file listing cache. Error is: {0}".format(files_in_path)
            self.send_thread_error(error_fmt)
            return error_fmt

        root_path_list = root_path.split("/")

        asset_info_sorted = dict()

        # go through all files under the root path
        for file_path in files_in_path:
            cgt_path = unicode(file_path["path"])
            # decompose the cgtw path into a list
            cgt_path_list = cgt_path.rstrip("/").split("/")

            # make sure path is not the root path
            if len(cgt_path_list) > len(root_path_list):

                # asset name for shot assets is different than show assets
                if asset_type is "shot":
                    asset_name = "{0}/{1}".format(
                        cgt_path_list[len(root_path_list)], cgt_path_list[len(root_path_list)+1]
                    )
                else:
                    asset_name = cgt_path_list[len(root_path_list)]

                # check if asset names were provided and if asset names provided check if the current asset is in list.
                # if not skip processing.
                if asset_names and asset_name not in asset_names:
                    continue

                # check if asset is published and goes through approval. Some asset components like rigs get approved
                # and have approved and work folders. Others like gpu caches don't.
                if self.is_asset_publishable(asset_type, asset_component):
                    # loop through folders that have asset information we want
                    for asset_folder in self.app_vars.asset_folder_list:
                        # check if path has one of the folders we want
                        if cgt_path.find("/" + asset_folder + "/") != -1:
                            # check if the asset dict has the asset, if not add
                            if asset_name not in asset_info_sorted:
                                asset_info_sorted[asset_name] = dict()

                            # check if the folder has been added, if not add
                            if asset_folder not in asset_info_sorted[asset_name]:
                                asset_info_sorted[asset_name][asset_folder] = list()

                            # add root_path for component
                            asset_info_sorted[asset_name]['component path'] = cgt_path.split(asset_folder)[0]

                            # add modified time
                            asset_info_sorted[asset_name]['modified time'] = file_path['modify_time']

                            # make sure we don't grab nested folders beneath the folders we want, first split
                            # at the folder
                            cgt_path_folder_parts = cgt_path.split(asset_folder)
                            # now split after the folder so we can see if there are nested folders
                            cgt_path_parts_after_history = cgt_path_folder_parts[1].split("/")
                            # there are nested folders if the count is > 2, so ignore. ie when split
                            # at '/', get ['','file] or ['','nested folder1',...]
                            if len(cgt_path_parts_after_history) <= 2:
                                # append file name
                                asset_info_sorted[asset_name][asset_folder].append(cgt_path)

                            # no need to continue processing, found one of the folders we want
                            break
                # asset isn't publishable
                else:
                    # split the component into a list if its multi-part like model/cache
                    component_list = asset_component.rstrip("/").split("/")
                    component_index = self._server_path_contains_asset_component(component_list, cgt_path_list)
                    # component found in path
                    if component_index:
                        start, end = component_index
                        if len(cgt_path_list) - end == 2:
                            # check for asset name
                            if asset_name not in asset_info_sorted and (asset_names is None or asset_name in asset_names):
                                asset_info_sorted[asset_name] = {'.': [cgt_path]}
                                # add root_path for component
                                asset_info_sorted[asset_name]['component path'] = '/'.join(cgt_path.split("/")[:-1])
                                # add modified time
                                asset_info_sorted[asset_name]['modified time'] = file_path['modify_time']
                            # asset exists
                            else:
                                # check if any files have been added under the asset name
                                if '.' not in asset_info_sorted[asset_name]:
                                    asset_info_sorted[asset_name]['.'] = []
                                asset_info_sorted[asset_name]['.'].append(cgt_path)

        return asset_info_sorted

    @staticmethod
    def _server_path_contains_asset_component(component_list, cgt_path):
        """
        Checks if a component is in the cgt path.
        :param component_list: This is a list, such as [rig] or [model, cache] since some components are multi-part
        like model cache
        :param cgt_path: the cgt server path as a string
        :return: the tuple of where the component starts in the path and ends in the path, or False if doesn't exist
        """
        # loop through path looking for component
        for i in xrange(1 + len(cgt_path) - len(component_list)):
            # if we find the component, return the start and end position
            if component_list == cgt_path[i:i+len(component_list)]:
                # return the start and end position of the component in the cgt path
                return i, i + len(component_list) - 1
        return False

    def _reset_thread_counters(self):
        # reset threads counters
        self.thread_total = 0.0
        self.threads_done = 0.0

    def _check_for_new_audio(self, seqs=None):
        """
        Call check_for_asset_updates and pass asset_component=audio to run this. Ensures proper polymorphism.
        Checks a sequence, or list of sequences, for any changed audio tracks. Checks all show sequences (using the
        sequences.json cache list in permanent data dir) if no sequence or sequences are provided. Uses mult-threading.
        Sequences with changed audio are stored in self.shots_with_changed_audio and any errors are stored in
        self.shots_failed_checking_timestamp.
        :param seqs: a list of sequences as a python list, where a sequence is Seq###. Also accepts strings. If
        providing multiple sequences in a string, separate with a comma. ex: "Seq040,Seq050"
        :return: Nothing
        """

        self.shots_failed_checking_timestamp = dict()
        self.shots_with_changed_audio = dict()

        # convert a string sequence name(s) to a list
        if not isinstance(seqs, list) and seqs:
            # check if a list of sequences was provided
            sequences = seqs.split(",")
            # check if multiple sequences provided
            if isinstance(sequences, list):
                # remove any spaces
                seqs = [seq.strip(" ") for seq in sequences]
            # just one sequence provided
            else:
                seqs = [seqs]

        # load the sequence/shot list
        error = self.ani_vars.load_seq_shot_list()
        if error:
            self.send_thread_error(error)
            return error

        # no sequences given, load all
        if not seqs:
            seqs = self.ani_vars.get_sequence_list()

        json_temp_file_info_path = os.path.join(
            self.app_vars.cgt_temp_file_cache_dir,
            "shot_audio_" + self.app_vars.cgt_tmp_file_cache_filename
        )

        # get file info for assets from CGT
        error = self.server_get_file_listing_using_folder_filter(
            "/LongGong/sequences", "audio", json_temp_file_info_path
        )
        if error:
            error_fmt = "Error getting file information from cgt server. Error is {0}".format(error)
            self.send_thread_error(error_fmt)
            return error_fmt
        # convert to asset list, ie sequence shot list
        audio_server_file_info = self._convert_cgt_file_info_to_asset_info(
            "/LongGong/sequences",
            json_temp_file_info_path,
            "shot",
            "audio"
        )

        # set number of threads to max - can do this since running per asset
        self.set_number_of_concurrent_threads()
        self._reset_thread_counters()

        # if not visible then no other function called this, so we can show progress window
        if not self.progress_win.isVisible():
            # reset progress
            self.init_progress_window("Audio Check Progress", "Checking all audio for changes...")

        # check audio for all sequences
        for seq in seqs:
            # create here, don't create in threads because could get race condition
            self.shots_failed_checking_timestamp[seq] = dict()
            self.shots_with_changed_audio[seq] = list()

            self.ani_vars.update(seq)
            for shot in self.ani_vars.get_shot_list():
                self.ani_vars.update(seq, shot)
                worker = pyani.core.ui.Worker(
                    self._check_shot_audio_timestamp,
                    False,
                    seq,
                    shot,
                    audio_server_file_info,
                    self.ani_vars.shot_audio_dir
                )
                self.thread_total += 1.0
                self.thread_pool.start(worker)
                # reset error list
                self.init_thread_error()
                # slot that is called when a thread finishes
                worker.signals.finished.connect(self._thread_audio_timestamp_check_complete)

    def _check_shot_audio_timestamp(self, seq, shot, audio_server_file_info, audio_dir):
        """
        This is a separate method so checking for audio changes can be threaded.
        Checks a shot to see if its audio changed. The first time a shot is checked, a file is stored locally with the
        server modified date. This starts the process of tracking this shot. The next time this method runs it will
        load the local file, and check the server modified date. If the server date is newer, that date is saved to
        the local file and the shot is added to the list of shots with changed audio
        :param seq: the sequence name, as Seq###
        :param shot: the shot name, as Shot###
        :param audio_server_file_info: a list of dicts, where each dict is a shot containing the audio's file info
        :param audio_dir: the local directory where audio is stored
        :return: Does not return a value, instead stores shots with changed audio in member variable
        shots_with_changed_audio. Errors are stored in member variable shots_failed_checking_timestamp as
        dict element in format shot name: 'error msg'
        """
        # variable to indicate if this is the first time to start tracking audio for a shot. Don't want to add the
        # shot to the report of changed audio if its the first time we are starting to track
        started_tracking = False

        # set default local date modified, in case the file doesn't exist
        local_audio_modified_date = datetime.strptime("2000-01-01 00:00:00", "%Y-%m-%d %H:%M:%S")
        # try to load audio info file which has last modified date, if doesn't exist, create, first time
        # checking this shot for audio
        audio_metadata_path = "{0}\\{1}".format(audio_dir, self.app_vars.audio_metadata_json_name)
        audio_metadata = pyani.core.util.load_json(audio_metadata_path)

        # if can't load, add last modified to file
        if not isinstance(audio_metadata, dict):
            audio_metadata = {"last_modified": ""}
        # file loaded, get the date time
        else:
            # check if the key exists in metadata
            if 'last_modified' not in audio_metadata:
                audio_metadata['last_modified'] = ""
            else:
                date_str = audio_metadata['last_modified']
                try:
                    local_audio_modified_date = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                # couldn't convert date time into date object
                except ValueError as error:
                    self.shots_failed_checking_timestamp[seq][shot] = error
                    return

        try:
            # get the server modified time for audio file
            asset_name = "{0}/{1}".format(seq, shot)
            modified_time = audio_server_file_info[asset_name]['modified time']
            server_modified_date = datetime.strptime(modified_time, "%Y-%m-%d %H:%M:%S")
        except KeyError:
            # key error so shot didn't have audio since no modify time
            error = "Shot does not have an audio file."
            self.shots_failed_checking_timestamp[seq][shot] = error
            return
        except ValueError:
            # couldn't convert date string to datetime object
            error = "Could not convert date from CGT server to a datetime object"
            self.shots_failed_checking_timestamp[seq][shot] = error
            return

        # if audio file is newer, save and update modified date in audio info file
        if server_modified_date > local_audio_modified_date:
            # write the server date time to disk and save shot for report
            timestamp = server_modified_date.strftime("%Y-%m-%d %H:%M:%S")
            audio_metadata['last_modified'] = timestamp

            # check if directory exists - because this check doesn't care if the audio file exists locally, just the
            # audio info file with the timestamp, the directory may not be there
            if not os.path.exists(audio_metadata_path):
                error = pyani.core.util.make_all_dir_in_path(audio_dir)
                if error:
                    self.shots_failed_checking_timestamp[seq][shot] = error
                # just created for the first time, flag it
                started_tracking = True

            error = pyani.core.util.write_json(audio_metadata_path, audio_metadata)
            # if an error occurred saving to disk, skip shot, otherwise it was written successfully so we
            # can record the shot as changed
            if error:
                self.shots_failed_checking_timestamp[seq][shot] = error
            # don't want to add newly tracked shots to report, i.e. skip them since we don't have any local date
            # information
            elif not started_tracking:
                self.shots_with_changed_audio[seq].append((shot, timestamp))

    def _thread_audio_timestamp_check_complete(self):
        """
        Called when a thread that checks an audio's timestamp completes
        """
        # a thread finished, increment our count
        self.threads_done += 1.0
        if self.threads_done > self.thread_total:
            return
        else:
            # get the current progress percentage
            progress = (self.threads_done / self.thread_total) * 100.0

            self.progress_win.setValue(progress)
            # check if we are finished
            if progress >= 100.0:
                # create excel report
                filename, error = self._generate_report_for_changed_audio()
                if error:
                    self.send_thread_error(error)
                    return error
                # done, let any listening objects/classes know we are finished, pass "audio" in case listeners want
                # to know what is sending this signal
                self.finished_tracking.emit(("audio", filename))

    def _generate_report_for_changed_audio(self):
        """
        Generates an excel workbook with a report of the shots with changed audio. Also reports any shots that errored
        while trying to find if audio changed
        :return: tuple of filename and error. error is None if created, otherwise error as string if not created
        """
        workbook = Workbook()
        active_sheet = workbook.active
        # start row
        row_index = 1
        # how many cells to merge
        merge_length = 20
        # to alternate shot row color between white and gray
        alternate_row_color = False
        # filename for saving workbook
        filename = "{0}\\{1}_{2}{3}".format(
            self.app_vars.audio_excel_report_dir,
            self.app_vars.audio_excel_report_filename,
            datetime.now().strftime("%B_%d_%H_%M"),
            self.app_vars.excel_ext
        )

        # setup styles
        seq_row_fill = PatternFill(patternType='solid', fgColor=Color(rgb="0099cccc"))
        shot_row_fill_shaded_light = PatternFill(patternType='solid', fgColor=Color(rgb="00f5f5f5"))
        shot_row_fill_shaded_dark = PatternFill(patternType='solid', fgColor=Color(rgb="00cccccc"))
        shot_row_fill_error = PatternFill(patternType='solid', fgColor=Color(rgb="00da9694"))
        seq_font_size = Font(size=20)
        shot_font_size = Font(size=12)

        # create data
        for seq in sorted(self.shots_with_changed_audio):
            active_sheet.cell(row_index, column=1).value = seq
            # apply fill color and font to cell
            active_sheet.cell(row_index, column=1).fill = seq_row_fill
            active_sheet.cell(row_index, column=1).font = seq_font_size
            # merge cells
            active_sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=merge_length)
            row_index += 1

            # a seq in shots_with_changed_audio may be empty, so check with a try..except so zip doesn't error
            # unpacking
            try:
                # a list of tuples so unpack into dict key value pair
                shot_date_list = dict(self.shots_with_changed_audio[seq])
                shot_list = shot_date_list.keys()
            except ValueError:
                shot_list = list()

            # does seq exist in errored list, if so get shots
            if seq in self.shots_failed_checking_timestamp:
                # get a list of all shots both errored and changed audio so they can be in order sorted
                shot_list.extend(self.shots_failed_checking_timestamp[seq].keys())

            for shot_index, shot in enumerate(sorted(shot_list)):
                if pyani.core.util.find_val_in_nested_dict(self.shots_failed_checking_timestamp, [seq, shot]):
                    active_sheet.cell(row_index, column=1).value = "{0} : {1}".format(
                        shot,
                        self.shots_failed_checking_timestamp[seq][shot]
                    )
                else:
                    active_sheet.cell(row_index, column=1).value = shot
                    # add space between shot name and date
                    active_sheet.cell(row_index, column=2).value = "        "
                    active_sheet.cell(row_index, column=3).value = shot_date_list[shot]

                # apply font size - don't care about column 2, its empty
                active_sheet.cell(row_index, column=1).font = shot_font_size
                active_sheet.cell(row_index, column=3).font = shot_font_size

                # figure out fill - error, or light or dark color and merge cells
                if pyani.core.util.find_val_in_nested_dict(self.shots_failed_checking_timestamp, [seq, shot]):
                    active_sheet.cell(row_index, column=1).fill = shot_row_fill_error
                    # merge cells
                    active_sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index,
                                             end_column=merge_length)
                elif alternate_row_color:
                    active_sheet.cell(row_index, column=1).fill = shot_row_fill_shaded_light
                    active_sheet.cell(row_index, column=2).fill = shot_row_fill_shaded_light
                    active_sheet.cell(row_index, column=3).fill = shot_row_fill_shaded_light
                    # merge cells
                    active_sheet.merge_cells(start_row=row_index, start_column=3, end_row=row_index,
                                             end_column=merge_length)
                else:
                    active_sheet.cell(row_index, column=1).fill = shot_row_fill_shaded_dark
                    active_sheet.cell(row_index, column=2).fill = shot_row_fill_shaded_dark
                    active_sheet.cell(row_index, column=3).fill = shot_row_fill_shaded_dark
                    # merge cells
                    active_sheet.merge_cells(start_row=row_index, start_column=3, end_row=row_index,
                                             end_column=merge_length)

                # flip the color for next row
                alternate_row_color = not alternate_row_color
                row_index += 1

        # save report
        if not os.path.exists(self.app_vars.audio_excel_report_dir):
            error = pyani.core.util.make_all_dir_in_path(self.app_vars.audio_excel_report_dir)
            if error:
                error_fmt = "Error creating directory for audio reports. Error is {0}".format(error)
                return None, error_fmt
        workbook.save(filename)

        # remove any old reports - get file listing and check if older than max days to keep
        for existing_report in os.listdir(self.app_vars.audio_excel_report_dir):
            existing_report_path = os.path.join(self.app_vars.audio_excel_report_dir, existing_report)
            error = pyani.core.util.delete_by_day(self.app_vars.audio_max_report_history, existing_report_path)
            if error:
                error_fmt = "Error removing old reports. Error is {0}".format(error)
                return None, error_fmt

        return filename, None
